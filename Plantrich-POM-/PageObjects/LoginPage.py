import openpyxl

class Login:
    path = "/home/coolminds/Desktop/PlantrichAutomationDatasheet.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet1 = workbook["Plantrich1"]

    USERNAME_XPATH="/html/body/app-root/app-login/div/main/div/div/div/div/div/div/form/div[2]/input"
    PASSWORD_XPATH="/html/body/app-root/app-login/div/main/div/div/div/div/div/div/form/div[3]/input"
    LOGIN_BUTTON_XPATH="/html/body/app-root/app-login/div/main/div/div/div/div/div/div/form/div[4]/div[1]/button"
    VENDOR_LOGIN_LINK_XPATH="/html/body/app-root/app-login/div/main/div/div/div/div/div/div/form/div[4]/div[2]/button"
    VENDOR_USERNAME_XPATH ="/html/body/app-root/app-vendor-login/div/main/div/div/div/div/div/div/form/div[1]/input"
    VENDOR_PASSWORD_XPATH ="/html/body/app-root/app-vendor-login/div/main/div/div/div/div/div/div/form/div[2]/input"
    VENDOR_LOGIN_BUTTON_XPATH="/html/body/app-root/app-vendor-login/div/main/div/div/div/div/div/div/form/div[3]/div[1]/button"
    FORGETPASSWORD_BUTTON_XPATH="/html/body/app-root/app-login/div/main/div/div/div/div/div/div/form/div[5]/div[2]/button"

    def __init__(self,driver):
        self.driver=driver

    def login(self):
        self.driver.find_element_by_xpath(self.USERNAME_XPATH).clear()
        self.driver.find_element_by_xpath(self.USERNAME_XPATH).send_keys(self.sheet1.cell(2,1).value)
        self.driver.find_element_by_xpath(self.PASSWORD_XPATH).clear()
        self.driver.find_element_by_xpath(self.PASSWORD_XPATH).send_keys(self.sheet1.cell(2,2).value)
        self.driver.find_element_by_xpath(self.LOGIN_BUTTON_XPATH).click()

    def Fake_login(self,username,password):
        self.driver.find_element_by_xpath(self.USERNAME_XPATH).clear()
        self.driver.find_element_by_xpath(self.USERNAME_XPATH).send_keys(username)
        self.driver.find_element_by_xpath(self.PASSWORD_XPATH).clear()
        self.driver.find_element_by_xpath(self.PASSWORD_XPATH).send_keys(password)
        self.driver.find_element_by_xpath(self.LOGIN_BUTTON_XPATH).click()


    def vendorlogin(self):
        self.driver.find_element_by_xpath(self.VENDOR_USERNAME_XPATH).clear()
        self.driver.find_element_by_xpath(self.VENDOR_USERNAME_XPATH).clear()
        self.driver.find_element_by_xpath(self.VENDOR_USERNAME_XPATH).send_keys(self.sheet1.cell(3,1).value)
        self.driver.find_element_by_xpath(self.VENDOR_PASSWORD_XPATH).clear()
        self.driver.find_element_by_xpath(self.VENDOR_PASSWORD_XPATH).send_keys(self.sheet1.cell(3,2).value)
        self.driver.find_element_by_xpath(self.VENDOR_LOGIN_BUTTON_XPATH).click()

    def Fake_vendorlogin(self,username,password):
        self.driver.find_element_by_xpath(self.VENDOR_USERNAME_XPATH).clear()
        self.driver.find_element_by_xpath(self.VENDOR_USERNAME_XPATH).clear()
        self.driver.find_element_by_xpath(self.VENDOR_USERNAME_XPATH).send_keys(username)
        self.driver.find_element_by_xpath(self.VENDOR_PASSWORD_XPATH).clear()
        self.driver.find_element_by_xpath(self.VENDOR_PASSWORD_XPATH).send_keys(password)
        self.driver.find_element_by_xpath(self.VENDOR_LOGIN_BUTTON_XPATH).click()

    def Forgetpassword(self):
        self.driver.find_element_by_xpath(self.FORGETPASSWORD_BUTTON_XPATH)

